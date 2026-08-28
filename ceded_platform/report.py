"""Build a clean, presentation-ready Excel report from a run.

One workbook: a Summary sheet up front (the answer at a glance), then formatted
detail sheets — styled headers, frozen panes, money formatting, auto widths.
"""

from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .sqlstore import reconciliations_frame

TEAL, DARK, GREEN, GREENBG, RED, REDBG = \
    "0C6E78", "0A565E", "2C6E4A", "DBEBE0", "B23A3A", "F4DADA"
HEADER_FILL = PatternFill("solid", fgColor=TEAL)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=18, color=DARK)
SUB_FONT = Font(size=10.5, color="586873")
SEC_FONT = Font(bold=True, size=12, color=DARK)
MONEY = "#,##0.00"
MONEY_KEYS = ("premium", "written", "earned", "unearned", "debit", "credit",
              "amount", "seeded", "allocated", "expected", "actual", "variance",
              "commission", "ibnr", "dac", "fee", "payable", "brokerage",
              "reserve", "collateral", "settlement", "fet")


def _excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    out = df.copy()
    for c in out.columns:
        if out[c].map(lambda v: isinstance(v, (list, dict))).any():
            out[c] = out[c].astype(str)
    return out


def _is_money(header) -> bool:
    return any(k in str(header).lower() for k in MONEY_KEYS)


def _style_data_sheet(ws) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    for col in ws.columns:
        idx = col[0].column
        header = ws.cell(row=1, column=idx).value
        width = max([len(str(header or ""))]
                    + [len(str(c.value)) for c in col[1:] if c.value is not None]
                    + [8])
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 3, 44)
        if _is_money(header):
            for c in col[1:]:
                if isinstance(c.value, (int, float)):
                    c.number_format = MONEY


def _section(ws, r, title):
    c = ws.cell(row=r, column=1, value=title)
    c.font = SEC_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    return r + 1


def _row(ws, r, values, header=False, money=(), bold=False):
    for i, v in enumerate(values):
        c = ws.cell(row=r, column=1 + i, value=v)
        if header:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        else:
            if bold:
                c.font = Font(bold=True)
            if i in money and isinstance(v, (int, float)):
                c.number_format = MONEY
    return r + 1


def _num(x) -> float:
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _build_summary(ws, res, recon, cfg=None) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    for c in "BCDE":
        ws.column_dimensions[c].width = 16

    ws["A1"] = "Ceded Reinsurance — Monthly Result"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    period = getattr(cfg, "current_period", "") if cfg else ""
    ws["A2"] = ("How much premium we cede to each reinsurer this period"
                + (f"  ·  {period}" if period else "")
                + "  —  with proof it ties out.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:E2")

    r = 4
    r = _section(ws, r, "1 · Ceded premium this month, by program")
    r = _row(ws, r, ["Program (Ceded ID)", "Ceded Written",
                     "Earned", "Unearned"], header=True)
    mv = res.movement
    if mv is not None and not mv.empty:
        for _, x in mv.iterrows():
            r = _row(ws, r, [x.get("Ceded ID"),
                             _num(x.get("Ceded Written Premium")),
                             _num(x.get("Ceded Earned Premium")),
                             _num(x.get("Ceded Unearned Premium"))],
                     money=[1, 2, 3])
        r = _row(ws, r, ["Total",
                         _num(mv["Ceded Written Premium"].sum()),
                         _num(mv["Ceded Earned Premium"].sum()),
                         _num(mv["Ceded Unearned Premium"].sum())],
                 money=[1, 2, 3], bold=True)
    r += 2

    r = _section(ws, r, "2 · Split across reinsurers")
    r = _row(ws, r, ["Reinsurer", "Ceded Written Premium"], header=True)
    books = getattr(res, "reinsurer_books", {}) or {}
    if books:
        for name, bk in books.items():
            tot = _num(bk["Reinsurer Ceded Written Premium"].sum()) \
                if "Reinsurer Ceded Written Premium" in bk.columns else 0.0
            r = _row(ws, r, [name, tot], money=[1])
    else:
        r = _row(ws, r, ["(no reinsurer panel for these programs)", 0.0])
    r += 2

    r = _section(ws, r, "3 · Validation — a human can verify this")
    npass = int(recon["passed"].sum()) if not recon.empty else 0
    n = len(recon)
    allp = (npass == n and n > 0)
    c = ws.cell(row=r, column=1,
                value=f"{npass} of {n} reconciliation checks passed"
                      + ("  ✓" if allp else "  ✗"))
    c.font = Font(bold=True, size=12, color=GREEN if allp else RED)
    c.fill = PatternFill("solid", fgColor=GREENBG if allp else REDBG)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    je = res.journal_entry
    bal = (abs(_num(je["Debit"].sum()) - _num(je["Credit"].sum())) < 0.01) \
        if (je is not None and not je.empty) else True
    ws.cell(row=r, column=1,
            value="Journal entry balances (debits = credits)  ✓" if bal
            else "Journal entry does NOT balance  ✗").font = \
        Font(size=11, color=GREEN if bal else RED)
    r += 2
    ws.cell(row=r, column=1,
            value="See the tabs for detail: Movement, Journal Entry, "
                  "Allocation, Reconciliations, and the inputs.").font = SUB_FONT


def build_report(res, inputs: dict, path, cfg=None):
    """Write the presentation-ready workbook to `path`."""
    recon = reconciliations_frame(res.reconciliations)
    sheets = {
        "Movement": res.movement,
        "Journal Entry": res.journal_entry,
        "Allocation": res.allocation_audit,
        "Reconciliations": recon,
        "ITD Pivot": res.itd_pivot,
        "Detail": res.fact_ceded_calc,
        "Input QS": inputs.get("qs"),
        "Input FAC": inputs.get("fac"),
        "Input EB": inputs.get("eb"),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        pd.DataFrame({" ": [" "]}).to_excel(xw, "Summary", index=False)  # sheet 1
        for name, df in sheets.items():
            _excel_safe(df if df is not None else pd.DataFrame()).to_excel(
                xw, sheet_name=name, index=False)
        wb = xw.book
        for ws in wb.worksheets:
            if ws.title != "Summary":
                _style_data_sheet(ws)
        s = wb["Summary"]
        s.delete_rows(1, s.max_row)
        _build_summary(s, res, recon, cfg)
    return path
